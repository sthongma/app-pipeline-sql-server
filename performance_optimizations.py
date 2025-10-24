"""
Performance optimization utilities for PIPELINE_SQLSERVER.

Provides optimized file processing capabilities:
1. Chunked file reading for large datasets
2. Parallel processing support
3. Enhanced memory management
4. Detailed progress tracking
5. Cancellation support
"""

import gc
import logging
import os
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd


class PerformanceOptimizer:
    """Handles performance optimization for file processing operations."""
    
    def __init__(self, log_callback: Optional[Callable[[str], None]] = None) -> None:
        """
        Initialize performance optimizer.
        
        Args:
            log_callback: Callback function for logging messages
        """
        self.log_callback = log_callback or logging.info
        self.cancellation_token = threading.Event()
        self.chunk_size = 50000  # Default optimized chunk size for large files
        self.max_workers = min(4, os.cpu_count() or 1)  # Number of worker threads
        
    def set_cancellation_token(self, token: threading.Event) -> None:
        """Set cancellation token for operation cancellation."""
        self.cancellation_token = token
        
    def get_optimal_chunk_size(self, file_size_mb: float) -> int:
        """Calculate optimal chunk size based on file size and available memory."""
        if file_size_mb < 50:
            return 25000
        elif file_size_mb < 100:
            return 50000
        elif file_size_mb < 200:
            return 75000
        elif file_size_mb < 500:
            return 100000
        else:
            return 150000  # For very large files
        
    def read_large_file_chunked(self, file_path: str, file_type: str = 'excel') -> Tuple[bool, pd.DataFrame]:
        """
        Read large files using chunked approach to save memory.
        
        Args:
            file_path: Path to the file
            file_type: File type ('excel', 'excel_xls', or 'csv')
            
        Returns:
            Tuple[bool, pd.DataFrame]: (success status, resulting DataFrame)
        """
        try:
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            
            self.log_callback(f"📁 Read File: {os.path.basename(file_path)} ({file_size_mb:.1f} MB)")
            
            # Set optimal chunk size based on file size
            optimal_chunk_size = self.get_optimal_chunk_size(file_size_mb)
            if optimal_chunk_size != self.chunk_size:
                self.chunk_size = optimal_chunk_size
                self.log_callback(f"🔧 Optimized chunk size for this file: {self.chunk_size:,} rows")
            
            if file_size_mb > 50:  # Lower threshold for chunked reading
                self.log_callback(f"⚠️ Large File, Use Chunked Reading")
                return self._read_large_file_chunked(file_path, file_type)
            else:
                return self._read_small_file(file_path, file_type)
                
        except Exception as e:
            error_msg = f"❌ Error Reading File: {e}"
            self.log_callback(error_msg)
            return False, pd.DataFrame()
    
    def _read_small_file(self, file_path: str, file_type: str) -> Tuple[bool, pd.DataFrame]:
        """Read small files using standard approach."""
        try:
            if file_type == 'csv':
                df = pd.read_csv(file_path, header=0, encoding='utf-8', dtype=str)
            elif file_type == 'excel_xls':
                df = pd.read_excel(file_path, header=0, sheet_name=0, engine='xlrd', dtype=str)
            else:
                df = pd.read_excel(file_path, header=0, sheet_name=0, engine='openpyxl', dtype=str)

            self.log_callback(f"✅ Read File Success: {len(df):,} rows, {len(df.columns)} columns")
            return True, df
            
        except Exception as e:
            self.log_callback(f"❌ Error Reading File: {e}")
            return False, pd.DataFrame()
    
    def _read_large_file_chunked(self, file_path: str, file_type: str) -> Tuple[bool, pd.DataFrame]:
        """Read large files using chunked approach."""
        try:
            chunks = []
            total_rows = 0
            
            if file_type == 'csv':
                total_rows, encoding_used = self._get_csv_info(file_path)
                self.log_callback(f"📊 Total Rows: {total_rows:,} (encoding={encoding_used})")
                
                # Read in chunks with proper encoding
                chunks = self._read_csv_chunks(file_path, encoding_used)
                        
            elif file_type == 'excel_xls':
                chunks = self._read_xls_chunks(file_path)
                        
            else:  # Excel .xlsx file
                chunks = self._read_xlsx_chunks(file_path)
            
            # Combine chunks with memory optimization
            if chunks:
                self.log_callback(f"🔄 Combining {len(chunks)} chunks...")
                
                # Progressive concatenation to avoid memory spike
                if len(chunks) > 10:  # For large number of chunks
                    self.log_callback("💾 Using progressive concatenation for large dataset")
                    result = chunks[0]
                    for i, chunk in enumerate(chunks[1:], 1):
                        if self.cancellation_token.is_set():
                            self.log_callback("❌ Work Cancelled")
                            break
                        
                        result = pd.concat([result, chunk], ignore_index=True)
                        
                        # Progress feedback
                        if i % 5 == 0:
                            progress = (i / (len(chunks) - 1)) * 100
                            self.log_callback(f"🔗 Combining: {i}/{len(chunks)-1} chunks ({progress:.1f}%)")
                        
                        # Memory cleanup
                        del chunk
                        if i % 5 == 0:
                            gc.collect()
                    
                    df = result
                else:
                    # Standard concatenation for smaller datasets
                    df = pd.concat(chunks, ignore_index=True)
                
                del chunks  # Release memory
                gc.collect()
                
                self.log_callback(f"✅ Read File Success - {len(df):,} rows, {len(df.columns)} columns")
                return True, df
            else:
                self.log_callback("❌ No data in file")
                return False, pd.DataFrame()
                
        except Exception as e:
            self.log_callback(f"❌ Error Reading File: {e}")
            return False, pd.DataFrame()
    
    def _get_csv_info(self, file_path: str) -> Tuple[int, str]:
        """Get CSV file information including encoding and row count."""
        # Try UTF-8 first, then fallback to cp874/latin1 for legacy files
        for encoding in ['utf-8', 'cp874', 'latin1']:
            try:
                total_rows = sum(1 for _ in open(file_path, 'r', encoding=encoding)) - 1
                return total_rows, encoding
            except UnicodeDecodeError:
                continue
        return 0, 'utf-8'  # Default fallback
    
    def _read_csv_chunks(self, file_path: str, encoding: str) -> List[pd.DataFrame]:
        """Read CSV file in chunks with optimized performance."""
        chunks = []

        import warnings
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=pd.errors.DtypeWarning)
            # Use optimized chunk size and memory settings
            chunk_reader = pd.read_csv(file_path, header=0, encoding=encoding,
                                     chunksize=self.chunk_size, low_memory=False,
                                     engine='c', dtype=str)  # Use C engine for better performance, read as string

        self.log_callback("💡 Using optimized CSV reader with C engine")
        
        total_processed = 0
        for i, chunk in enumerate(chunk_reader):
            if self.cancellation_token.is_set():
                self.log_callback("❌ Work Cancelled")
                break
            
            chunks.append(chunk)
            total_processed += len(chunk)
            
            # Enhanced progress feedback
            self.log_callback(f"📖 Chunk {i+1}: {len(chunk):,} rows (Total: {total_processed:,})")
            
            # Aggressive memory cleanup for large files
            if (i + 1) % 5 == 0:
                gc.collect()
                self.log_callback(f"🧹 Memory cleanup after {i+1} chunks")
        
        return chunks
    
    def _read_xls_chunks(self, file_path: str) -> List[pd.DataFrame]:
        """Read XLS file in chunks."""
        import xlrd

        chunks = []
        workbook = xlrd.open_workbook(file_path)
        worksheet = workbook.sheet_by_index(0)

        # Read headers
        headers = [str(worksheet.cell_value(0, col_idx)) for col_idx in range(worksheet.ncols)]

        # Read data in chunks
        chunk_data = []
        for row_idx in range(1, worksheet.nrows):
            if self.cancellation_token.is_set():
                self.log_callback("❌ Work Cancelled")
                break

            # Convert all cell values to string to prevent scientific notation
            row_data = [str(worksheet.cell_value(row_idx, col_idx)) for col_idx in range(worksheet.ncols)]
            chunk_data.append(row_data)

            # Create chunk every chunk_size rows
            if len(chunk_data) >= self.chunk_size:
                chunk_df = pd.DataFrame(chunk_data, columns=headers)
                chunks.append(chunk_df)
                chunk_data = []

                self.log_callback(f"📖 Read Chunk {len(chunks)}: {len(chunk_df):,} rows")
                gc.collect()

        # Add remaining data
        if chunk_data:
            chunk_df = pd.DataFrame(chunk_data, columns=headers)
            chunks.append(chunk_df)

        return chunks
    
    def _read_xlsx_chunks(self, file_path: str) -> List[pd.DataFrame]:
        """Read XLSX file in chunks with optimized performance."""
        import openpyxl
        
        chunks = []
        self.log_callback("🔄 Opening Excel file with read-only mode...")
        
        # Use read-only mode with data_only for better performance
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        
        # Get total rows for progress tracking
        total_rows = worksheet.max_row - 1  # Exclude header
        self.log_callback(f"📊 Total rows to process: {total_rows:,}")
        
        # Read headers using optimized method
        headers = [str(cell.value) if cell.value is not None else '' for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

        # Read data in larger chunks with batch processing
        chunk_data = []
        processed_rows = 0

        # Use iter_rows for better performance instead of cell-by-cell access
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if self.cancellation_token.is_set():
                self.log_callback("❌ Work Cancelled")
                workbook.close()
                break

            # Convert all cell values to string to prevent scientific notation
            chunk_data.append([str(cell) if cell is not None else '' for cell in row])
            processed_rows += 1
            
            # Progress feedback every 10,000 rows
            if processed_rows % 10000 == 0:
                progress = (processed_rows / total_rows) * 100
                self.log_callback(f"📖 Processing: {processed_rows:,}/{total_rows:,} rows ({progress:.1f}%)")
            
            # Create chunk when reaching chunk_size
            if len(chunk_data) >= self.chunk_size:
                chunk_df = pd.DataFrame(chunk_data, columns=headers)
                chunks.append(chunk_df)
                chunk_data = []
                
                chunk_num = len(chunks)
                self.log_callback(f"✅ Completed Chunk {chunk_num}: {len(chunk_df):,} rows")
                
                # Aggressive memory cleanup for large files
                del chunk_df
                gc.collect()
        
        # Add remaining data
        if chunk_data:
            chunk_df = pd.DataFrame(chunk_data, columns=headers)
            chunks.append(chunk_df)
            self.log_callback(f"✅ Final Chunk {len(chunks)}: {len(chunk_df):,} rows")
        
        workbook.close()
        self.log_callback(f"🎯 Chunking Complete: {len(chunks)} chunks created")
        return chunks
    
    def process_dataframe_in_chunks(self, df: pd.DataFrame, chunk_size: int = 5000) -> List[pd.DataFrame]:
        """
        Split DataFrame into chunks for processing.
        
        Args:
            df: DataFrame to split
            chunk_size: Size of each chunk
            
        Returns:
            List[pd.DataFrame]: List of DataFrame chunks
        """
        chunks = []
        total_rows = len(df)
        
        for i in range(0, total_rows, chunk_size):
            end_idx = min(i + chunk_size, total_rows)
            chunk = df.iloc[i:end_idx].copy()
            chunks.append(chunk)
            
            if self.cancellation_token.is_set():
                break
        
        return chunks
    
    def parallel_process_files(self, file_paths: List[str], process_func: Callable, 
                             progress_callback: Optional[Callable] = None) -> List[Tuple[bool, Any]]:
        """
        Process multiple files in parallel.
        
        Args:
            file_paths: List of file paths
            process_func: Function to process each file
            progress_callback: Callback for progress updates
            
        Returns:
            List[Tuple[bool, Any]]: Processing results
        """
        results = []
        completed = 0
        total_files = len(file_paths)
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit jobs to executor
            future_to_file = {
                executor.submit(process_func, file_path): file_path 
                for file_path in file_paths
            }
            
            # Wait for results
            for future in as_completed(future_to_file):
                if self.cancellation_token.is_set():
                    self.log_callback("❌ Work Cancelled")
                    break
                
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    results.append((True, result))
                except Exception as e:
                    self.log_callback(f"❌ Error Processing File: {os.path.basename(file_path)}: {e}")
                    results.append((False, str(e)))
                
                completed += 1
                if progress_callback:
                    progress = completed / total_files
                    progress_callback(progress, f"Processing File {completed}/{total_files}")
        
        return results
    
    def optimize_memory_usage(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Optimize DataFrame memory usage.
        
        Args:
            df: DataFrame to optimize
            
        Returns:
            pd.DataFrame: Optimized DataFrame
        """
        try:
            initial_memory = df.memory_usage(deep=True).sum() / 1024 / 1024  # MB
            self.log_callback(f"💾 Memory Start: {initial_memory:.2f} MB")
            
            # Downcast numeric columns
            for col in df.select_dtypes(include=['int64']).columns:
                df[col] = pd.to_numeric(df[col], downcast='integer')
            
            for col in df.select_dtypes(include=['float64']).columns:
                df[col] = pd.to_numeric(df[col], downcast='float')
            
            # Optimize object columns
            for col in df.select_dtypes(include=['object']).columns:
                if df[col].nunique() / len(df) < 0.5:  # If unique values < 50%
                    df[col] = df[col].astype('category')
            
            final_memory = df.memory_usage(deep=True).sum() / 1024 / 1024  # MB
            memory_saved = initial_memory - final_memory
            
            self.log_callback(f"💾 Memory After Optimization: {final_memory:.2f} MB (Saved {memory_saved:.2f} MB)")
            
            return df
            
        except Exception as e:
            self.log_callback(f"⚠️ Cannot optimize memory: {e}")
            return df
    
    def create_progress_tracker(self, total_items: int, description: str = "") -> Callable:
        """
        Create progress tracker for monitoring task progress.
        
        Args:
            total_items: Total number of items to process
            description: Description of the operation
            
        Returns:
            Callable: Function to update progress
        """
        start_time = time.time()
        completed = 0
        
        def update_progress(items_completed: int = 1, custom_message: str = "") -> Tuple[float, str]:
            nonlocal completed
            completed += items_completed
            
            if total_items > 0:
                progress = completed / total_items
                elapsed_time = time.time() - start_time
                
                if completed > 0:
                    estimated_total = elapsed_time / completed * total_items
                    remaining_time = estimated_total - elapsed_time
                    
                    message = custom_message or f"{description}: {completed}/{total_items}"
                    time_info = f" (remaining {remaining_time:.1f}s)"
                    
                    return progress, message + time_info
                else:
                    return 0.0, f"{description}: Start..."
            else:
                return 1.0, f"{description}: Finished"
        
        return update_progress
    
    def cleanup_memory(self) -> None:
        """Clean up memory by forcing garbage collection."""
        gc.collect()
        self.log_callback("🧹 Cleaned up memory")


class LargeFileProcessor:
    """Specialized processor for handling large files."""
    
    def __init__(self, log_callback: Optional[Callable[[str], None]] = None) -> None:
        """
        Initialize large file processor.
        
        Args:
            log_callback: Callback function for logging messages
        """
        self.optimizer = PerformanceOptimizer(log_callback)
        self.log_callback = log_callback or logging.info
        
    def process_large_file(self, file_path: str, file_type: str, 
                          processing_steps: List[Callable[[pd.DataFrame], pd.DataFrame]]) -> Tuple[bool, pd.DataFrame]:
        """
        Process large file with defined processing steps.
        
        Args:
            file_path: Path to the file
            file_type: Type of the file
            processing_steps: List of processing functions
            
        Returns:
            Tuple[bool, pd.DataFrame]: (success status, processed DataFrame)
        """
        try:
            # Step 1: Read file
            success, df = self.optimizer.read_large_file_chunked(file_path, file_type)
            if not success:
                return False, pd.DataFrame()
            
            # Step 2: Optimize memory
            df = self.optimizer.optimize_memory_usage(df)
            
            # Step 3: Process with defined steps
            for i, step_func in enumerate(processing_steps):
                if self.optimizer.cancellation_token.is_set():
                    self.log_callback("❌ Work Cancelled")
                    return False, pd.DataFrame()
                
                self.log_callback(f"🔄 Step {i+1}/{len(processing_steps)}: {step_func.__name__}")
                df = step_func(df)
                
                # Clean memory after each step
                self.optimizer.cleanup_memory()
            
            return True, df
            
        except Exception as e:
            self.log_callback(f"❌ Error Processing File: {e}")
            return False, pd.DataFrame()
    
    def set_cancellation_token(self, token: threading.Event) -> None:
        """Set cancellation token for operation cancellation."""
        self.optimizer.set_cancellation_token(token)


# Helper functions for processing
def create_chunk_processor(chunk_size: int = 5000) -> Callable[[pd.DataFrame, Callable], pd.DataFrame]:
    """Create function for chunk-based data processing."""
    def process_in_chunks(df: pd.DataFrame, process_func: Callable[[pd.DataFrame], pd.DataFrame]) -> pd.DataFrame:
        """Process DataFrame in chunks."""
        results = []
        total_chunks = (len(df) + chunk_size - 1) // chunk_size
        
        for i in range(0, len(df), chunk_size):
            chunk = df.iloc[i:i+chunk_size].copy()
            processed_chunk = process_func(chunk)
            results.append(processed_chunk)
            
            # Show progress
            chunk_num = (i // chunk_size) + 1
            logging.info(f"📊 Processing chunk {chunk_num}/{total_chunks}")
        
        return pd.concat(results, ignore_index=True)
    
    return process_in_chunks


def estimate_processing_time(file_size_mb: float, processing_type: str = 'standard') -> float:
    """
    Estimate processing time based on file size.
    
    Args:
        file_size_mb: File size in MB
        processing_type: Type of processing
        
    Returns:
        float: Estimated time in seconds
    """
    # Approximate processing rates (MB/second)
    rates = {
        'fast': 10.0,      # 10 MB/second
        'standard': 5.0,   # 5 MB/second
        'slow': 2.0        # 2 MB/second
    }
    
    rate = rates.get(processing_type, rates['standard'])
    estimated_time = file_size_mb / rate
    
    return estimated_time


def format_file_size(size_bytes: int) -> str:
    """Convert file size to human-readable format."""
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.1f} {size_names[i]}"


def format_time(seconds: float) -> str:
    """Convert time to human-readable format."""
    if seconds < 60:
        return f"{seconds:.1f} seconds"
    elif seconds < 3600:
        minutes = seconds / 60
        return f"{minutes:.1f} minutes"
    else:
        hours = seconds / 3600
        return f"{hours:.1f} hours"